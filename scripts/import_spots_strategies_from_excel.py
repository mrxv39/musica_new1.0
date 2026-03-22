from __future__ import annotations

import argparse
import math
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from modules.db.conn import connect
from modules.db.migrate import init_db


@dataclass(frozen=True)
class ImportCounts:
    scopes: int
    or_rules: int
    nash: int


def _norm_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _as_float_or_none(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        x = float(v)
        if math.isnan(x):
            return None
        return x
    s = _norm_text(v)
    if s == "":
        return None
    # allow comma decimal
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def _header_map(ws) -> dict[str, int]:
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    return {_norm_text(v): i + 1 for i, v in enumerate(header) if _norm_text(v) != ""}


def _iter_rows(ws) -> Iterable[int]:
    for r in range(2, ws.max_row + 1):
        yield r


def _validate_stack_bounds(row_idx: int, se_min: float | None, se_max: float | None) -> None:
    if se_min is None or se_max is None:
        raise ValueError(f"Row {row_idx}: missing stack_effective_min/max")
    if se_min > se_max:
        raise ValueError(f"Row {row_idx}: stack_effective_min > stack_effective_max ({se_min} > {se_max})")


def _validate_bet_bounds(row_idx: int, bet_min: float | None, bet_max: float | None) -> None:
    if bet_min is None or bet_max is None:
        return
    if bet_min > bet_max:
        raise ValueError(f"Row {row_idx}: bet_min > bet_max ({bet_min} > {bet_max})")


def import_spots_strategies_from_excel(xlsx_path: Path, *, wipe: bool) -> ImportCounts:
    if not xlsx_path.exists():
        raise FileNotFoundError(str(xlsx_path))

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    scopes_ws = wb["Strategy scopes"]
    hoja1_ws = wb["Hoja1"]

    scopes_h = _header_map(scopes_ws)
    hoja1_h = _header_map(hoja1_ws)

    # Discover all Nash-type sheets from scopes (any sheet_name != "Hoja1")
    nash_sheet_names = set()
    for r in _iter_rows(scopes_ws):
        sn = _norm_text(scopes_ws.cell(r, scopes_h["sheet_name"]).value)
        if sn and sn != "Hoja1":
            nash_sheet_names.add(sn)

    init_db()
    with connect() as conn:
        cur = conn.cursor()

        if wipe:
            cur.execute("DELETE FROM spots_strategy_scopes")
            cur.execute("DELETE FROM spots_strategies")
            cur.execute("DELETE FROM spots_strategies_nash")

        scopes_count = 0
        for r in _iter_rows(scopes_ws):
            spot_key = _norm_text(scopes_ws.cell(r, scopes_h["spot_key"]).value)
            if spot_key == "":
                continue
            p2_tipo = _norm_text(scopes_ws.cell(r, scopes_h["p2_tipo"]).value)
            p3_tipo = _norm_text(scopes_ws.cell(r, scopes_h["p3_tipo"]).value)
            se_min = _as_float_or_none(scopes_ws.cell(r, scopes_h["scope_se_min"]).value)
            se_max = _as_float_or_none(scopes_ws.cell(r, scopes_h["scope_se_max"]).value)
            sheet_name = _norm_text(scopes_ws.cell(r, scopes_h["sheet_name"]).value)
            _validate_stack_bounds(r, se_min, se_max)
            if sheet_name == "":
                raise ValueError(f"Row {r}: missing sheet_name")

            cur.execute(
                """
                INSERT INTO spots_strategy_scopes (
                    spot_key, p2_tipo, p3_tipo, scope_se_min, scope_se_max, sheet_name
                ) VALUES (?, ?, ?, ?, ?, ?)
                """,
                (spot_key, p2_tipo, p3_tipo, se_min, se_max, sheet_name),
            )
            scopes_count += 1

        or_count = 0
        for r in _iter_rows(hoja1_ws):
            spot_key = _norm_text(hoja1_ws.cell(r, hoja1_h["spot_key"]).value)
            if spot_key == "":
                continue
            hand_range_name = _norm_text(hoja1_ws.cell(r, hoja1_h["hand_range_name"]).value)
            move = _norm_text(hoja1_ws.cell(r, hoja1_h["move"]).value)
            bet_min_cell = hoja1_ws.cell(r, hoja1_h["bet_min"]).value
            bet_max_cell = hoja1_ws.cell(r, hoja1_h["bet_max"]).value
            bet_min_num = _as_float_or_none(bet_min_cell)
            bet_max_num = _as_float_or_none(bet_max_cell)
            hand_range = _norm_text(hoja1_ws.cell(r, hoja1_h["hand_range"]).value)
            se_min = _as_float_or_none(hoja1_ws.cell(r, hoja1_h["stack_effective_min"]).value)
            se_max = _as_float_or_none(hoja1_ws.cell(r, hoja1_h["stack_effective_max"]).value)
            _validate_stack_bounds(r, se_min, se_max)
            _validate_bet_bounds(r, bet_min_num, bet_max_num)

            def t(key: str) -> str:
                if key not in hoja1_h:
                    return ""
                return _norm_text(hoja1_ws.cell(r, hoja1_h[key]).value)

            def f(key: str) -> float | None:
                if key not in hoja1_h:
                    return None
                return _as_float_or_none(hoja1_ws.cell(r, hoja1_h[key]).value)

            cur.execute(
                """
                INSERT INTO spots_strategies (
                    spot_key, hand_range_name, move, bet_min, bet_max, hand_range,
                    stack_effective_min, stack_effective_max,
                    p1_pos, p2_pos, p3_pos, p2_tipo, p3_tipo,
                    p1bet_min, p1bet_max, p2bet_min, p2bet_max, p3bet_min, p3bet_max,
                    p2stack_min, p2stack_max, p3stack_min, p3stack_max
                ) VALUES (
                    ?, ?, ?, ?, ?, ?,
                    ?, ?,
                    ?, ?, ?, ?, ?,
                    ?, ?, ?, ?, ?, ?,
                    ?, ?, ?, ?
                )
                """,
                (
                    spot_key,
                    hand_range_name,
                    move,
                    bet_min_num if bet_min_num is not None else (_norm_text(bet_min_cell) or None),
                    bet_max_num if bet_max_num is not None else (_norm_text(bet_max_cell) or None),
                    hand_range,
                    se_min,
                    se_max,
                    t("p1_pos"),
                    t("p2_pos"),
                    t("p3_pos"),
                    t("p2_tipo"),
                    t("p3_tipo"),
                    f("p1bet_min"),
                    f("p1bet_max"),
                    f("p2bet_min"),
                    f("p2bet_max"),
                    f("p3bet_min"),
                    f("p3bet_max"),
                    f("p2stack_min"),
                    f("p2stack_max"),
                    f("p3stack_min"),
                    f("p3stack_max"),
                ),
            )
            or_count += 1

        nash_count = 0
        for nash_sheet_name in sorted(nash_sheet_names):
            if nash_sheet_name not in wb.sheetnames:
                print(f"WARNING: Sheet '{nash_sheet_name}' referenced in scopes but not found — skipped")
                continue
            nash_ws = wb[nash_sheet_name]
            nash_h = _header_map(nash_ws)
            for r in _iter_rows(nash_ws):
                spot_key = _norm_text(nash_ws.cell(r, nash_h["spot_key"]).value)
                if spot_key == "":
                    continue
                hand_range_name = _norm_text(nash_ws.cell(r, nash_h["hand_range_name"]).value)
                move = _norm_text(nash_ws.cell(r, nash_h["move"]).value)
                hand_class = _norm_text(nash_ws.cell(r, nash_h["hand_class"]).value)
                se_min = _as_float_or_none(nash_ws.cell(r, nash_h["stack_effective_min"]).value)
                se_max = _as_float_or_none(nash_ws.cell(r, nash_h["stack_effective_max"]).value)
                _validate_stack_bounds(r, se_min, se_max)
                if hand_class == "":
                    raise ValueError(f"Sheet '{nash_sheet_name}' Row {r}: missing hand_class")

                def t(key: str, _ws=nash_ws, _h=nash_h, _r=r) -> str:
                    if key not in _h:
                        return ""
                    return _norm_text(_ws.cell(_r, _h[key]).value)

                def f(key: str, _ws=nash_ws, _h=nash_h, _r=r) -> float | None:
                    if key not in _h:
                        return None
                    return _as_float_or_none(_ws.cell(_r, _h[key]).value)

                cur.execute(
                    """
                    INSERT INTO spots_strategies_nash (
                        spot_key, hand_range_name, move, hand_class,
                        stack_effective_min, stack_effective_max,
                        p1_pos, p2_pos, p3_pos, p2_tipo, p3_tipo,
                        p1bet_min, p1bet_max, p2bet_min, p2bet_max, p3bet_min, p3bet_max,
                        p2stack_min, p2stack_max, p3stack_min, p3stack_max
                    ) VALUES (
                        ?, ?, ?, ?,
                        ?, ?,
                        ?, ?, ?, ?, ?,
                        ?, ?, ?, ?, ?, ?,
                        ?, ?, ?, ?
                    )
                    """,
                    (
                        spot_key,
                        hand_range_name,
                        move,
                        hand_class,
                        se_min,
                        se_max,
                        t("p1_pos"),
                        t("p2_pos"),
                        t("p3_pos"),
                        t("p2_tipo"),
                        t("p3_tipo"),
                        f("p1bet_min"),
                        f("p1bet_max"),
                        f("p2bet_min"),
                        f("p2bet_max"),
                        f("p3bet_min"),
                        f("p3bet_max"),
                        f("p2stack_min"),
                        f("p2stack_max"),
                        f("p3stack_min"),
                        f("p3stack_max"),
                    ),
                )
                nash_count += 1

        conn.commit()

    return ImportCounts(scopes=scopes_count, or_rules=or_count, nash=nash_count)


def main() -> int:
    parser = argparse.ArgumentParser(description="Import spot strategies from Excel into SQLite DB.")
    parser.add_argument(
        "--xlsx",
        default=str(Path("data/spots_strategies.xlsx")),
        help="Path to spots_strategies.xlsx",
    )
    parser.add_argument(
        "--wipe",
        action="store_true",
        help="Delete existing rows before importing.",
    )
    args = parser.parse_args()

    counts = import_spots_strategies_from_excel(Path(args.xlsx), wipe=bool(args.wipe))
    print(f"Imported scopes={counts.scopes} or_rules={counts.or_rules} nash={counts.nash}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

